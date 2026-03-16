"""
api/excel_export.py

Builds a formatted Excel workbook with two tabs:

  Triangle  — original claims triangle + age-to-age LDF triangle (with Excel formulas
               referencing the data triangle above) + weighted-average LDF and CDF rows.

  Results   — three-method IBNR comparison (CL / BF / Cape Cod) where IBNR and
               loss-ratio columns are live Excel formulas, not hardcoded values.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
_DARK_BLUE  = "1F4E79"
_MED_BLUE   = "2E75B6"
_LIGHT_BLUE = "D6E4F0"
_ALT_ROW    = "EBF3FB"
_EMPTY_CELL = "ECECEC"
_TOTAL_FILL = "FFE699"
_WHITE      = "FFFFFF"
_BLACK      = "000000"

# ── Dev-period metadata (matches chain_ladder.py) ─────────────────────────────
_DEV_COLS    = ["dev_12", "dev_24", "dev_36", "dev_48", "dev_60", "dev_72"]
_DEV_PERIODS = [12, 24, 36, 48, 60, 72]
_LDF_KEYS    = [f"{_DEV_PERIODS[i]}→{_DEV_PERIODS[i+1]}" for i in range(len(_DEV_PERIODS) - 1)]
_LDF_LABELS  = ["12→24", "24→36", "36→48", "48→60", "60→72"]
_DEV_LABELS  = ["Dev 12", "Dev 24", "Dev 36", "Dev 48", "Dev 60", "Dev 72"]

N_DEV = len(_DEV_COLS)   # 6
N_LDF = len(_LDF_KEYS)   # 5


# ── Style helpers ──────────────────────────────────────────────────────────────

def _font(bold=False, size=11, color=_BLACK):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=False)

def _thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border(color=_MED_BLUE):
    return Border(bottom=Side(border_style="medium", color=color))


def _w(ws, row: int, col: int, value,
       bold=False, color=_BLACK, bg=None,
       h_align="right", fmt=None, border=None, size=11):
    """Write a value/formula to (row, col) with optional styling."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=size, color=color)
    c.alignment = _align(h=h_align)
    if bg:
        c.fill = _fill(bg)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


def _header_row(ws, row: int, labels: list[str], bg=_MED_BLUE,
                start_col: int = 1, row_height: int = 18):
    """Write a styled header row."""
    for j, lbl in enumerate(labels):
        _w(ws, row, start_col + j, lbl,
           bold=True, color=_WHITE, bg=bg, h_align="center")
    ws.row_dimensions[row].height = row_height


def _section_title(ws, row: int, title: str,
                   start_col: int, end_col: int, bg=_DARK_BLUE):
    """Write a merged section title."""
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    _w(ws, row, start_col, title,
       bold=True, color=_WHITE, bg=bg, h_align="center", size=12)
    ws.row_dimensions[row].height = 22


# ── Triangle tab ───────────────────────────────────────────────────────────────

def _build_triangle_tab(ws, triangle: pd.DataFrame, cl) -> None:
    years = list(triangle.index)
    n     = len(years)

    # Column widths
    ws.column_dimensions["A"].width = 17
    for c in range(2, N_DEV + 2):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws.freeze_panes = "B3"

    # ── Section 1: Claims Triangle ────────────────────────────────────────────
    TITLE_ROW = 1
    HDR_ROW   = 2
    D0        = 3                           # first data row

    _section_title(ws, TITLE_ROW, "Claims Development Triangle",
                   start_col=1, end_col=N_DEV + 1)
    _header_row(ws, HDR_ROW, ["Accident Year"] + _DEV_LABELS)

    for i, year in enumerate(years):
        row = D0 + i
        bg  = _ALT_ROW if i % 2 else None
        _w(ws, row, 1, year, bold=True, bg=bg, h_align="center")
        for j, col_name in enumerate(_DEV_COLS):
            val = triangle.loc[year, col_name]
            if pd.notna(val):
                _w(ws, row, 2 + j, int(val), bg=bg, h_align="right", fmt="#,##0")
            else:
                _w(ws, row, 2 + j, None, bg=_EMPTY_CELL)

    # ── Section 2: LDF Triangle ───────────────────────────────────────────────
    LDF_TITLE = D0 + n + 1          # blank gap at D0+n
    LDF_HDR   = LDF_TITLE + 1
    LDF_D0    = LDF_HDR + 1
    LDF_GAP   = LDF_D0 + n          # blank gap after LDF data
    AVG_ROW   = LDF_GAP + 1
    CDF_ROW   = AVG_ROW + 1

    _section_title(ws, LDF_TITLE, "Age-to-Age Link Ratio (LDF) Triangle",
                   start_col=1, end_col=N_LDF + 1)
    _header_row(ws, LDF_HDR, ["Accident Year"] + _LDF_LABELS)

    # LDF data: Excel formulas referencing the data triangle above
    for i, year in enumerate(years):
        ldf_row  = LDF_D0 + i
        data_row = D0 + i           # the corresponding row in the data triangle
        bg       = _ALT_ROW if i % 2 else None
        _w(ws, ldf_row, 1, year, bold=True, bg=bg, h_align="center")
        for j in range(N_LDF):
            from_name = _DEV_COLS[j]
            to_name   = _DEV_COLS[j + 1]
            if pd.notna(triangle.loc[year, from_name]) and pd.notna(triangle.loc[year, to_name]):
                from_letter = get_column_letter(2 + j)
                to_letter   = get_column_letter(2 + j + 1)
                formula = f"={to_letter}{data_row}/{from_letter}{data_row}"
                _w(ws, ldf_row, 2 + j, formula, bg=bg, h_align="right", fmt="0.0000")
            else:
                _w(ws, ldf_row, 2 + j, None, bg=_EMPTY_CELL)

    # Volume-weighted average LDF row (actual CL values — no formula shortcut since
    # they aggregate across different AY rows; show the computed values directly)
    _w(ws, AVG_ROW, 1, "Volume-Weighted Avg LDF",
       bold=True, bg=_LIGHT_BLUE, h_align="left")
    for j, key in enumerate(_LDF_KEYS):
        _w(ws, AVG_ROW, 2 + j, round(float(cl.ldfs[key]), 4),
           bold=True, bg=_LIGHT_BLUE, h_align="right", fmt="0.0000",
           border=_bottom_border())

    # CDF to Ultimate row — spans all 6 dev periods (B-G)
    _w(ws, CDF_ROW, 1, "CDF to Ultimate",
       bold=True, bg=_LIGHT_BLUE, h_align="left")
    for j, dev_col in enumerate(_DEV_COLS):
        _w(ws, CDF_ROW, 2 + j, round(float(cl.cdfs[dev_col]), 4),
           bold=True, bg=_LIGHT_BLUE, h_align="right", fmt="0.0000")


# ── Results tab ────────────────────────────────────────────────────────────────

_RESULT_HEADERS = [
    "Accident Year", "Paid to Date", "Premium",
    "CL Ultimate",   "BF Ultimate",  "CC Ultimate",
    "CL IBNR",       "BF IBNR",      "CC IBNR",
    "CL Loss Ratio", "BF Loss Ratio","CC Loss Ratio",
]

_COL_WIDTHS = [15, 15, 15, 15, 15, 15, 14, 14, 14, 14, 14, 14]


def _build_results_tab(ws, triangle: pd.DataFrame, bf, cc,
                       fitted_tail_factor: float) -> None:
    years  = list(triangle.index)
    n      = len(years)
    bf_df  = bf.comparison
    cc_sum = cc.cc_summary

    # Column widths
    for col, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "B3"

    TITLE_ROW = 1
    HDR_ROW   = 2
    D0        = 3
    D_END     = D0 + n - 1
    TOT_ROW   = D_END + 2           # one blank row gap
    ASSUMP_HDR = TOT_ROW + 2

    # Title
    _section_title(ws, TITLE_ROW, "Reserve Estimates by Accident Year",
                   start_col=1, end_col=12)

    # Column headers
    _header_row(ws, HDR_ROW, _RESULT_HEADERS)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, year in enumerate(years):
        row    = D0 + i
        bg     = _ALT_ROW if i % 2 else None
        bf_row = bf_df.loc[year]
        cc_row = cc_sum.loc[year]

        paid    = int(bf_row["paid_to_date"])
        premium = int(bf_row["premium"])
        cl_ult  = int(bf_row["cl_ultimate"])
        bf_ult  = int(bf_row["bf_ultimate"])
        cc_ult  = int(cc_row["cc_ultimate"])

        # Columns A–F: hardcoded values
        _w(ws, row, 1,  year,    bold=True, bg=bg, h_align="center")
        _w(ws, row, 2,  paid,    bg=bg, fmt="#,##0")
        _w(ws, row, 3,  premium, bg=bg, fmt="#,##0")
        _w(ws, row, 4,  cl_ult,  bg=bg, fmt="#,##0")
        _w(ws, row, 5,  bf_ult,  bg=bg, fmt="#,##0")
        _w(ws, row, 6,  cc_ult,  bg=bg, fmt="#,##0")

        # Columns G–I: IBNR = Ultimate − Paid to Date  (Excel formulas)
        _w(ws, row, 7,  f"=D{row}-B{row}", bg=bg, fmt="#,##0")   # CL IBNR
        _w(ws, row, 8,  f"=E{row}-B{row}", bg=bg, fmt="#,##0")   # BF IBNR
        _w(ws, row, 9,  f"=F{row}-B{row}", bg=bg, fmt="#,##0")   # CC IBNR

        # Columns J–L: Loss Ratio = Ultimate / Premium  (Excel formulas)
        _w(ws, row, 10, f"=D{row}/C{row}", bg=bg, fmt="0.00%")   # CL LR
        _w(ws, row, 11, f"=E{row}/C{row}", bg=bg, fmt="0.00%")   # BF LR
        _w(ws, row, 12, f"=F{row}/C{row}", bg=bg, fmt="0.00%")   # CC LR

    # ── Totals row ─────────────────────────────────────────────────────────────
    _w(ws, TOT_ROW, 1, "TOTAL",
       bold=True, bg=_TOTAL_FILL, h_align="center")

    # SUM formulas for currency columns B–I
    for col in range(2, 10):
        cl = get_column_letter(col)
        _w(ws, TOT_ROW, col,
           f"=SUM({cl}{D0}:{cl}{D_END})",
           bold=True, bg=_TOTAL_FILL, fmt="#,##0",
           border=_bottom_border(_DARK_BLUE))

    # Total loss ratios = sum(ultimates) / sum(premiums)  (Excel formulas)
    _w(ws, TOT_ROW, 10, f"=D{TOT_ROW}/C{TOT_ROW}",
       bold=True, bg=_TOTAL_FILL, fmt="0.00%",
       border=_bottom_border(_DARK_BLUE))
    _w(ws, TOT_ROW, 11, f"=E{TOT_ROW}/C{TOT_ROW}",
       bold=True, bg=_TOTAL_FILL, fmt="0.00%",
       border=_bottom_border(_DARK_BLUE))
    _w(ws, TOT_ROW, 12, f"=F{TOT_ROW}/C{TOT_ROW}",
       bold=True, bg=_TOTAL_FILL, fmt="0.00%",
       border=_bottom_border(_DARK_BLUE))

    # ── Assumptions section ────────────────────────────────────────────────────
    _section_title(ws, ASSUMP_HDR, "Model Assumptions",
                   start_col=1, end_col=4, bg=_MED_BLUE)

    assump_rows = [
        ("BF A Priori ELR",  bf.elr if isinstance(bf.elr, float) else "per-year",  "0.00%",
         "Derived CC ELR",   cc.cc_elr,                                             "0.00%"),
        ("Tail Factor Used", bf.tail_factor,                                         "0.0000",
         "Fitted Tail Factor", fitted_tail_factor,                                   "0.0000"),
    ]
    for idx, (lbl1, val1, fmt1, lbl2, val2, fmt2) in enumerate(assump_rows):
        r = ASSUMP_HDR + 1 + idx
        _w(ws, r, 1, lbl1, bold=True, bg=_LIGHT_BLUE, h_align="left")
        _w(ws, r, 2, val1, bg=_ALT_ROW, fmt=fmt1)
        _w(ws, r, 3, lbl2, bold=True, bg=_LIGHT_BLUE, h_align="left")
        _w(ws, r, 4, val2, bg=_ALT_ROW, fmt=fmt2)


# ── Public entry point ─────────────────────────────────────────────────────────

def build_excel(
    triangle: pd.DataFrame,
    cl,
    bf,
    cc,
    fitted_tail_factor: float,
) -> bytes:
    """
    Build a two-tab Excel workbook and return its contents as bytes.

    Parameters
    ----------
    triangle           : cleaned dev triangle (from load_triangle)
    cl                 : fitted ChainLadder (run() already called; fit_tail() already called)
    bf                 : fitted BornhuetterFerguson (run() already called)
    cc                 : fitted CapeCod (run() already called)
    fitted_tail_factor : result of cl.fit_tail(), surfaced in the assumptions section

    Returns
    -------
    bytes
        Raw .xlsx bytes suitable for streaming in a FastAPI response.
    """
    wb = Workbook()

    # Rename the default sheet and create a second one
    ws_tri = wb.active
    ws_tri.title = "Triangle"
    ws_res = wb.create_sheet("Results")

    _build_triangle_tab(ws_tri, triangle, cl)
    _build_results_tab(ws_res, triangle, bf, cc, fitted_tail_factor)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
